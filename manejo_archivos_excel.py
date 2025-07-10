#¿Que es openpyxl?
#Es una libreria para trabajar con archivos excel

#Crear un archivo de Excel 

from openpyxl import load_workbook,Workbook

#Importar los estilos
from openpyxl.styles import Font

workbook = Workbook()
workbook.save("./Python-Basico/relatos.xlsx")

#seleccionarlo y agregarle valores

book = workbook.active

#Escribir datos

book["A1"] = "Producto"
book["B1"] = "Precio"
book["C1"] = "Cantidad"

#Valores

book["A2"] = "Camisa"
book["B2"] = "5000"
book["C2"] = "10"

#Para guardar cambios utilizamos el metodo save()

workbook.save("./Python-Basico/relatos.xlsx")

#para leer tendre la clase load_workbook

#Aqui solo cargo el archivo sin abrirlo

book = load_workbook("./Python-Basico/relatos.xlsx")

hoja = book.active

#Vamos a leer los datos

product = hoja["A2"].value
price = hoja["B2"].value
quantity = hoja["C2"].value

print(f"Producto: {product}, Precio: {price}, Cantidad: {quantity}") 

#Crear un libro que se llama ventas y cargarlo

book = Workbook()

#Crear una hoja
hoja = book.active

#Cambiar el titulo

hoja.title = "Ventas-Enero"

#encabezados

hoja["A1"] = "Dia"
hoja["B1"] = "Ventas (USD)"

book.save("./Python-Basico/ventas.xlsx")

hoja["A1"].font = Font(bold=True)
hoja["B1"].font = Font(bold=True)

#creemos valores

#Un array con varios arrays dentro se llama matriz
ventas = [
  ["Lunes", 100],
  ["Martes", 200],
  ["Miercoles", 300],
  ["Jueves", 400],
  ["Viernes", 500],
  ["Sabado", 600],
  ["Domingo", 700],
]

#Para insertar mutiples valores necesitaremos utilizar un bucle

for row in ventas:
  hoja.append(row) #agregara este valor automaticamente a la siguiente fila vacia
  
book.save("./Python-Basico/ventas.xlsx")

#leer el contenido

book = load_workbook("./Python-Basico/ventas.xlsx")

hoja = book.active

#mostrar mi contenido

for row in hoja.iter_rows(min_row=2):
  #Cada row se ve asi (celda1=dia, celda2=ventas)
  print(row[0].value, row[1].value)
  
#Ejercicio 

#Crea un archivo ventas_trimestrales y agrega los datos de la siguiente manera:
# [trimestre, ventas totales]

#luego ingresa los valores de todos los trimestres de un año

#Luego muestra las ventas totales de cada trimestre



# Crear un nuevo libro de trabajo
workbook_trimestral = Workbook()
hoja_trimestral = workbook_trimestral.active
hoja_trimestral.title = "Ventas Trimestrales"

# Encabezados
hoja_trimestral["A1"] = "Trimestre"
hoja_trimestral["B1"] = "Ventas Totales"

# Aplicar estilo de negrita a los encabezados
hoja_trimestral["A1"].font = Font(bold=True)
hoja_trimestral["B1"].font = Font(bold=True)

# Datos de ventas trimestrales
ventas_trimestrales = [
    ["Primer Trimestre", 15000],
    ["Segundo Trimestre", 22000],
    ["Tercer Trimestre", 18000],
    ["Cuarto Trimestre", 25000],
]

# Insertar los datos en la hoja
for row_data in ventas_trimestrales:
    hoja_trimestral.append(row_data)

# Guardar el archivo
workbook_trimestral.save("./Python-Basico/ventas_trimestrales.xlsx")

print("Archivo 'ventas_trimestrales.xlsx' creado exitosamente.")

# --- Leer y mostrar el contenido del archivo ---

# Cargar el archivo de Excel
book_loaded = load_workbook("./Python-Basico/ventas_trimestrales.xlsx")
hoja_loaded = book_loaded.active

print("\n--- Ventas Totales por Trimestre ---")
# Iterar sobre las filas, comenzando desde la segunda fila para omitir los encabezados
for row in hoja_loaded.iter_rows(min_row=2):
    trimestre = row[0].value
    ventas_totales = row[1].value
    print(f"{trimestre}: ${ventas_totales:,}")